import json
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class OS:
    def __init__(self, name: str, version: str):
        self.name = name
        self.version = version

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'OS':
        return cls(
            name=data['name'],
            version=data['version']
        )

class Software:
    def __init__(self, name: str, version: str, vendor: str):
        self.name = name
        self.version = version
        self.vendor = vendor

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'Software':
        return cls(
            name=data['name'],
            version=data['version'],
            vendor=data['vendor']
        )

class Vulnerability:
    def __init__(self, 
                 kaspersky_id: str, 
                 product_name: str, 
                 description_url: str, 
                 recommended_major_patch: str, 
                 recommended_minor_patch: str, 
                 severity_str: str, 
                 severity: int, 
                 cve: List[str], 
                 exploit_exists: bool, 
                 malware_exists: bool):
        self.kaspersky_id = kaspersky_id
        self.product_name = product_name
        self.description_url = description_url
        self.recommended_major_patch = recommended_major_patch
        self.recommended_minor_patch = recommended_minor_patch
        self.severity_str = severity_str
        self.severity = severity
        self.cve = cve
        self.exploit_exists = exploit_exists
        self.malware_exists = malware_exists

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'Vulnerability':
        return cls(
            kaspersky_id=data['kasperskyID'],
            product_name=data['productName'],
            description_url=data['descriptionURL'],
            recommended_major_patch=data['recommendedMajorPatch'],
            recommended_minor_patch=data['recommendedMinorPatch'],
            severity_str=data['severityStr'],
            severity=data['severity'],
            cve=data['cve'],
            exploit_exists=data['exploitExists'],
            malware_exists=data['malwareExists']
        )

class Device:
    def __init__(self, 
                 name: str, 
                 fqdn: List[str], 
                 ip_addresses: List[str], 
                 mac_addresses: List[str], 
                 owner: str, 
                 os: OS, 
                 software: List[Software], 
                 vulnerabilities: List[Vulnerability]):
        self.name = name
        self.fqdn = fqdn
        self.ip_addresses = ip_addresses
        self.mac_addresses = mac_addresses
        self.owner = owner
        self.os = os
        self.software = software
        self.vulnerabilities = vulnerabilities

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'Device':
        return cls(
            name=data['name'],
            fqdn=data['fqdn'],
            ip_addresses=data['ipAddresses'],
            mac_addresses=data['macAddresses'],
            owner=data['owner'],
            os=OS.from_dict(data['os']),
            software=[Software.from_dict(item) for item in data['software']],
            vulnerabilities=[Vulnerability.from_dict(item) for item in data['vulnerabilities']]
        )

def save_to_excel(devices: List[Device], output_file: str):
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create worksheets
    devices_ws = wb.create_sheet("Device")
    software_ws = wb.create_sheet("Software")
    vulnerabilities_ws = wb.create_sheet("Vulnerability")
    
    # Format styles
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    
    # Write Devices sheet
    devices_headers = [
        "Name", "FQDN", "IP Addresses", "MAC Addresses", "Owner", "OS Name", "OS Version"
    ]
    for col_num, header in enumerate(devices_headers, 1):
        cell = devices_ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    for row_num, device in enumerate(devices, 2):
        devices_ws.cell(row=row_num, column=1, value=device.name).alignment = wrap_alignment
        devices_ws.cell(row=row_num, column=2, value=", ".join(device.fqdn)).alignment = wrap_alignment
        devices_ws.cell(row=row_num, column=3, value=", ".join(device.ip_addresses)).alignment = wrap_alignment
        devices_ws.cell(row=row_num, column=4, value=", ".join(device.mac_addresses)).alignment = wrap_alignment
        devices_ws.cell(row=row_num, column=5, value=device.owner).alignment = wrap_alignment
        devices_ws.cell(row=row_num, column=6, value=device.os.name).alignment = wrap_alignment
        devices_ws.cell(row=row_num, column=7, value=device.os.version).alignment = wrap_alignment
        
        # Apply borders
        for col in range(1, 8):
            devices_ws.cell(row=row_num, column=col).border = thin_border
    
    # Write Software sheet with grouping by device name
    software_headers = [
        "��", "Software Name", "������", "������"
    ]
    for col_num, header in enumerate(software_headers, 1):
        cell = software_ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    row_num = 2
    for device in devices:
        start_row = row_num
        if not device.software:
            software_ws.cell(row=row_num, column=1, value=device.name).alignment = wrap_alignment
            for col in range(1, 5):
                software_ws.cell(row=row_num, column=col).border = thin_border
            row_num += 1
            continue
            
        for software in device.software:
            software_ws.cell(row=row_num, column=1, value=device.name if row_num == start_row else "").alignment = wrap_alignment
            software_ws.cell(row=row_num, column=2, value=software.name).alignment = wrap_alignment
            software_ws.cell(row=row_num, column=3, value=software.version).alignment = wrap_alignment
            software_ws.cell(row=row_num, column=4, value=software.vendor).alignment = wrap_alignment
            
            # Apply borders
            for col in range(1, 5):
                software_ws.cell(row=row_num, column=col).border = thin_border
            
            row_num += 1
        
        # Merge device name cells
        if row_num > start_row + 1:
            software_ws.merge_cells(start_row=start_row, start_column=1, end_row=row_num-1, end_column=1)
    
    # Write Vulnerabilities sheet with grouping by device name
    vuln_headers = [
        "��", "Kaspersky ID", "Product Name", "Severity", 
        "Severity Level", "CVE IDs", "Exploit Exists", "Malware Exists",
        "Recommended Major Patch", "Recommended Minor Patch", "Description URL"
    ]
    for col_num, header in enumerate(vuln_headers, 1):
        cell = vulnerabilities_ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    row_num = 2
    for device in devices:
        start_row = row_num
        if not device.vulnerabilities:
            vulnerabilities_ws.cell(row=row_num, column=1, value=device.name).alignment = wrap_alignment
            for col in range(1, 12):
                vulnerabilities_ws.cell(row=row_num, column=col).border = thin_border
            row_num += 1
            continue
            
        for vuln in device.vulnerabilities:
            vulnerabilities_ws.cell(row=row_num, column=1, value=device.name if row_num == start_row else "").alignment = wrap_alignment
            vulnerabilities_ws.cell(row=row_num, column=2, value=vuln.kaspersky_id).alignment = wrap_alignment
            vulnerabilities_ws.cell(row=row_num, column=3, value=vuln.product_name).alignment = wrap_alignment
            vulnerabilities_ws.cell(row=row_num, column=4, value=vuln.severity).alignment = wrap_alignment
            vulnerabilities_ws.cell(row=row_num, column=5, value=vuln.severity_str).alignment = wrap_alignment
            vulnerabilities_ws.cell(row=row_num, column=6, value=", ".join(vuln.cve)).alignment = wrap_alignment
            vulnerabilities_ws.cell(row=row_num, column=7, value="Yes" if vuln.exploit_exists else "No").alignment = wrap_alignment
            vulnerabilities_ws.cell(row=row_num, column=8, value="Yes" if vuln.malware_exists else "No").alignment = wrap_alignment
            vulnerabilities_ws.cell(row=row_num, column=9, value=vuln.recommended_major_patch).alignment = wrap_alignment
            vulnerabilities_ws.cell(row=row_num, column=10, value=vuln.recommended_minor_patch).alignment = wrap_alignment
            vulnerabilities_ws.cell(row=row_num, column=11, value=vuln.description_url).alignment = wrap_alignment
            
            # Apply borders
            for col in range(1, 12):
                vulnerabilities_ws.cell(row=row_num, column=col).border = thin_border
            
            row_num += 1
        
        # Merge device name cells
        if row_num > start_row + 1:
            vulnerabilities_ws.merge_cells(start_row=start_row, start_column=1, end_row=row_num-1, end_column=1)
    
    # Auto-adjust column widths with maximum limit (900px ~ 100 chars)
    MAX_COLUMN_WIDTH = 100
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        
        # Adjust column widths with maximum limit
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Calculate width with limit
            adjusted_width = min((max_length + 2) * 1.2, MAX_COLUMN_WIDTH)
            ws.column_dimensions[column].width = adjusted_width
        
        # Adjust row heights for wrapped text
        for row in ws.iter_rows():
            max_lines = 1
            for cell in row:
                if cell.value and "\n" in str(cell.value):
                    lines = len(str(cell.value).split("\n"))
                    if lines > max_lines:
                        max_lines = lines
            if max_lines > 1:
                ws.row_dimensions[row[0].row].height = 15 * max_lines
    
    # Save the workbook
    wb.save(output_file)
    print(f"Data successfully saved to {output_file}")

def main():
    try:
        with open('response.json', 'r', encoding='utf-8') as f:
            data = json.load(f)
    except FileNotFoundError:
        print("Error: File 'response.json' not found.")
        return
    except json.JSONDecodeError:
        print("Error: Invalid JSON format in 'response.json'.")
        return
    
    devices = [Device.from_dict(item) for item in data]
    save_to_excel(devices, "devices_grouped_rows.xlsx")
    print(f"Processed {len(devices)} devices. Results saved to devices_grouped_rows.xlsx")

if __name__ == "__main__":
    main()